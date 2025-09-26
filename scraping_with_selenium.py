#!/usr/bin/env python3
# scraping_with_selenium.py - Seleniumを使ったスクレイピング
#
# drivers/chromedriver.exeを使用しています。
#
# このスクリプトは、GitHub Pagesで公開された練習用の静的サイトを対象にしています。
#
# リンクをたどってページを取得していきます。



import os, openpyxl, time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment

driver_path = os.path.join("drivers", "chromedriver.exe")

url = 'https://nyaataco.github.io/scraping_practice_site/page1.html'

output_dir = 'output'
output_file = 'output_selenium.xlsx'

# Excelスタイル設定
fill_color = 'FDE9D9'
border_color = '542804'

title_font_style = Font(size=20, bold=True)
header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
header_font_size = Font(bold=True)
header_alignment = Alignment(horizontal='center', vertical='center')

header_border = Border(
    left=Side(style='thin', color=border_color),
    right=Side(style='thin', color=border_color),
    top=Side(style='medium', color=border_color),
    bottom=Side(style='thin', color=border_color)
)

side_border = Border(
    left=Side(style='thin', color=border_color),
    right=Side(style='thin', color=border_color),
)

last_row_border =  Border(
    left=Side(style='thin', color=border_color),
    right=Side(style='thin', color=border_color),
    bottom=Side(style='medium', color=border_color)
)


os.makedirs(output_dir, exist_ok=True)

# Excel 初期化
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'JobList'

sheet['A1'] = 'スクレイピングした情報一覧'
sheet['A1'].font = title_font_style
sheet.row_dimensions[1].height = 30
sheet.row_dimensions[3].height = 22

sheet['A3'] = '求人タイトル'
sheet['B3'] = '説明'
sheet['C3'] = '職種'
sheet['D3'] = '雇用形態'
sheet['E3'] = '給与'
sheet['F3'] = '企業名'
sheet['G3'] = '所在地'
sheet['H3'] = 'リンク'



# Selenium 起動
service = Service(driver_path)
browser = webdriver.Chrome(service=service)

try:
    count = 1
    has_next = True
    row = 4

    while has_next:
        # ページを取得
        print(f"ページを取得しています。: {url}")
        browser.get(url)
        time.sleep(1)

        cards = browser.find_elements(by="class name", value="jobsearch-card-link")
        for card in cards:
            link = card.get_attribute("href")
            status = card.find_element(by="class name", value="employment-status").text
            title = card.find_element(by="class name", value="jobsearch-title").text
            description = card.find_element(by="class name", value="description").text
            location = card.find_element(by="class name", value="jobsearch-location").text
            industry = card.find_element(by="class name", value="jobsearch-industry").text
            salary = card.find_element(by="class name", value="jobsearch-salary").text
            company = card.find_element(by="class name", value="jobsearch-company").text

            sheet[f'A{row}'] = title
            sheet[f'B{row}'] = description
            sheet[f'C{row}'] = industry
            sheet[f'D{row}'] = status
            sheet[f'E{row}'] = salary
            sheet[f'F{row}'] = company
            sheet[f'G{row}'] = location
            sheet[f'H{row}'] = link

            row += 1

        # 次のページへのリンクを探す
        next_url = None
        pagenators = browser.find_element(by="class name", value="pagenator")
        if pagenators:
            pages = pagenators.find_elements(By.TAG_NAME, "a")

            for page in pages:
                div = page.find_element(by="class name", value="pagenator-num")
                classes = div.get_attribute("class")

                if "current-page" in classes:
                    continue

                if int(page.text) > count:
                    next_url = page.get_attribute("href")
                    break

        if next_url:
            url = next_url
            count += 1
        else:
            has_next = False
            print("最後のページに到達しました。")


    # excelにスタイルをつける
    for cell in sheet[3]:
        cell.fill = header_fill
        cell.font = header_font_size
        cell.alignment = header_alignment

    sheet.column_dimensions['A'].width = 25 # 求人タイトル
    sheet.column_dimensions['B'].width = 30 # 説明
    sheet.column_dimensions['C'].width = 25 # 職種
    sheet.column_dimensions['D'].width = 10 # 雇用形態
    sheet.column_dimensions['E'].width = 25 # 給与
    sheet.column_dimensions['F'].width = 30 # 企業名
    sheet.column_dimensions['G'].width = 30 # 所在地
    sheet.column_dimensions['H'].width = 35 # リンク

    # 罫線
    for row_num in range(3, sheet.max_row + 1):
        sheet.row_dimensions[row_num].height = 20
        if row_num == 3:
            for col_num, cell in enumerate(sheet[row_num], start=1):
                if col_num == 1:
                    cell.border = Border(left=Side(style='medium', color=border_color), right=Side(style='thin', color=border_color), top=Side(style='medium', color=border_color), bottom=Side(style='thin', color=border_color))
                if col_num == len(sheet[row_num]):
                    cell.border = Border(left=Side(style='thin', color=border_color), right=Side(style='medium', color=border_color), top=Side(style='medium', color=border_color), bottom=Side(style='thin', color=border_color))
                else:
                    cell.border = header_border
        elif row_num == sheet.max_row:
            for col_num, cell in enumerate(sheet[row_num], start=1):
                if col_num == 1:
                    cell.border = Border(left=Side(style='medium', color=border_color), right=Side(style='thin', color=border_color), bottom=Side(style='medium', color=border_color))
                if col_num == len(sheet[row_num]):
                    cell.border = Border(left=Side(style='thin', color=border_color), right=Side(style='medium', color=border_color), bottom=Side(style='medium', color=border_color))
                else:
                    cell.border = last_row_border
        else:
            for col_num, cell in enumerate(sheet[row_num], start=1):
                if col_num == 1:
                    cell.border = Border(left=Side(style='medium', color=border_color), right=Side(style='thin', color=border_color),)
                if col_num == len(sheet[row_num]):
                    cell.border = Border(left=Side(style='thin', color=border_color), right=Side(style='medium', color=border_color),)
                else:
                    cell.border = side_border


    save_file = os.path.join(output_dir, output_file)
    wb.save(save_file)
    print(f'{save_file}を書き出しました。')

except Exception as e:
    print(f'処理を完了できませんでした。: {e}')

finally:
    browser.quit() 
